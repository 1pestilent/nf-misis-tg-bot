from openpyxl import load_workbook

import zipfile
import os

from time import sleep
from PIL import Image

from app.schedule.dirs import *
from app.schedule.downloader import downloader, timecalc
from app.schedule.handler import converter

def delete_3_sheet(path):
    workbook = load_workbook(path)
    if len(workbook.sheetnames) < 3:
        print(f"[Ошибка] В документе'{path}' нет нужного листа!")
        return path
    else:
        while len(workbook.sheetnames) >= 3:
            sheet_to_delete = workbook.sheetnames[2]
            workbook.remove(workbook[sheet_to_delete])
            workbook.save(path)
            print(f"[Успешно] Лист '{sheet_to_delete}' удалён из документа {path}.")
    


def unzip():
    
    files = os.listdir(zip_dir)

    for zip in files:
        if zip.endswith('.zip'):
            with zipfile.ZipFile(f'{zip_dir}/{zip}', 'r') as zipper:
                zipper.extractall(pdf_dir)
                print(f"[Успешно] Архив {zip} успешно разархивирован")




def get_schedule(course, week):
    path = downloader.download_schedule(course, week)
    if path:
        delete_3_sheet(path)
        pdf_path = converter.xlsx_to_pdf(path)
        pngs = converter.pdf_to_png(pdf_path, course, week)
        return pngs
    else:
        return False
    